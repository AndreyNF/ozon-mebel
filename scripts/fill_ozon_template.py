#!/usr/bin/env python3
"""Записать строку карточки Ozon в xlsx-шаблон «Комплекты мебели»."""

from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path

import openpyxl

TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "Комплекты мебели_21.05.2026.xlsx"
SHEET = "Шаблон"
HEADER_ROW = 2
FIRST_DATA_ROW = 5

# column index -> header name (row 2)
COLUMNS = {
    1: "№",
    2: "Артикул*",
    3: "Название товара",
    4: "Цена, руб.*",
    5: "Цена до скидки, руб.",
    6: "НДС, %*",
    7: "Рассрочка",
    8: "Ускоренный сбор отзывов",
    9: "SKU",
    10: "Штрихкод (Серийный номер / EAN)",
    11: "Вес в упаковке, г*",
    12: "Ширина упаковки, мм*",
    13: "Высота упаковки, мм*",
    14: "Длина упаковки, мм*",
    15: "Ссылка на главное фото*",
    16: "Ссылки на дополнительные фото",
    17: "Ссылки на фото 360",
    18: "Артикул фото",
    19: "Бренд*",
    20: "Название модели (для объединения в одну карточку)*",
    21: "Название цвета",
    22: "Цвет товара",
    23: "Тип*",
    24: "ТН ВЭД коды ЕАЭС*",
    25: "Количество товара в УЕИ",
    26: "Минимальное количество оптом",
    27: "#Хештеги",
    28: "Аннотация",
    29: "Rich-контент JSON",
    30: "Объединить в похожие товары",
    31: "Гарантийный срок",
    32: "Гарантия",
    33: "Состав комплекта",
    34: "Материал корпуса",
    35: "Вес товара, г",
    36: "Образец цвета",
    37: "Макс. нагрузка, кг",
    38: "Особенности",
    39: "Назначение (помещение)",
    40: "Комната",
    41: "Исполнение",
    42: "Стиль дизайна",
    43: "Форма поставки",
    44: "Количество заводских упаковок",
    45: "Вид выпуска товара",
    46: "Страна-изготовитель",
    47: "Комплектация",
}

HEADER_TO_COL = {v: k for k, v in COLUMNS.items()}
# normalize keys without asterisk for lookup
for col, name in list(COLUMNS.items()):
    HEADER_TO_COL[name.replace("*", "")] = col


def find_row_by_article(ws, article: str) -> int | None:
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val and str(val).strip() == article:
            return r
    return None


def find_empty_row(ws) -> int:
    for r in range(FIRST_DATA_ROW, ws.max_row + 100):
        val = ws.cell(r, 2).value
        if val is None or str(val).strip() == "":
            return r
    return ws.max_row + 1


def write_row(ws, row_index: int, data: dict) -> None:
    for key, value in data.items():
        if value is None or value == "":
            continue
        col = HEADER_TO_COL.get(key) or HEADER_TO_COL.get(key.replace("*", ""))
        if col:
            ws.cell(row_index, col, value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill Ozon furniture set template row")
    parser.add_argument("json_path", type=Path, help="JSON with field values keyed by header names")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output xlsx path (default: same dir as json / filled-template.xlsx)",
    )
    parser.add_argument(
        "-t",
        "--template",
        type=Path,
        default=TEMPLATE,
        help="Source template xlsx",
    )
    args = parser.parse_args()

    data = json.loads(args.json_path.read_text(encoding="utf-8"))
    article = data.get("Артикул*") or data.get("Артикул")
    if not article:
        raise SystemExit("JSON must contain Артикул* (manufacturer article)")

    out = args.output or args.json_path.parent / "filled-template.xlsx"
    if out.resolve() != args.template.resolve():
        shutil.copy2(args.template, out)

    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    row = find_row_by_article(ws, str(article)) or find_empty_row(ws)

    # photo URLs: accept images list
    if "images" in data and isinstance(data["images"], list):
        urls = [u for u in data["images"] if u]
        if urls:
            data.setdefault("Ссылка на главное фото*", urls[0])
            if len(urls) > 1:
                data.setdefault("Ссылки на дополнительные фото", "\n".join(urls[1:]))

    write_row(ws, row, data)
    wb.save(out)
    print(f"OK row {row} article {article} -> {out}")


if __name__ == "__main__":
    main()
