#!/usr/bin/env python3
"""Собрать финальный xlsx для загрузки в Ozon из row.json."""

from __future__ import annotations

import json
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "templates" / "Комплекты мебели_21.05.2026.xlsx"
SHEET = "Шаблон"
HEADER_ROW = 2
FIRST_DATA_ROW = 5

HEADER_TO_COL = {}


def load_headers(ws) -> dict[str, int]:
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(HEADER_ROW, c).value
        if h:
            mapping[str(h).strip()] = c
            mapping[str(h).replace("*", "").strip()] = c
    return mapping


def find_row(ws, article: str) -> int:
    for r in range(FIRST_DATA_ROW, ws.max_row + 200):
        val = ws.cell(r, 2).value
        if val and str(val).strip() == article:
            return r
        if val is None or str(val).strip() == "":
            return r
    return FIRST_DATA_ROW


def main() -> None:
    article = sys.argv[1] if len(sys.argv) > 1 else "Ц0081444"
    card = ROOT / "cards" / article
    row_path = card / f"{article}.row.json"
    row = json.loads(row_path.read_text(encoding="utf-8"))

    out = card / f"OZON_UPLOAD_{article}_{date.today().isoformat()}.xlsx"
    if out.exists():
        try:
            with open(out, "a"):
                pass
        except PermissionError:
            for n in range(2, 20):
                alt = card / f"OZON_UPLOAD_{article}_{date.today().isoformat()}_v{n}.xlsx"
                if not alt.exists():
                    out = alt
                    break
                try:
                    with open(alt, "a"):
                        pass
                    out = alt
                    break
                except PermissionError:
                    continue
    shutil.copy2(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    headers = load_headers(ws)
    target = find_row(ws, article)

    skip = {"images", "image_files", "_meta", "_packaging", "_photos"}
    photo_extra_key = "Ссылки на дополнительные фото"
    for key, value in row.items():
        if key.startswith("_") or key in skip:
            continue
        if value is None or value == "" or value == "[УТОЧНИТЬ]":
            continue
        if key == photo_extra_key and isinstance(value, str):
            # Ozon/Excel на Windows: каждый URL с новой строки (CRLF), как Alt+Enter в ячейке
            parts = [u.strip() for u in value.replace("\r", "").split("\n") if u.strip()]
            value = "\r\n".join(parts)
        col = headers.get(key) or headers.get(key.replace("*", ""))
        if col:
            cell = ws.cell(target, col, value)
            if key == photo_extra_key:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[target].height = max(120, 15 * len(parts))

    wb.save(out)
    print(str(out))

    try:
        import subprocess

        subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "card_registry.py"), "register", article, "--status", "ready"],
            cwd=ROOT,
            check=False,
            capture_output=True,
        )
    except Exception:
        pass


if __name__ == "__main__":
    main()
